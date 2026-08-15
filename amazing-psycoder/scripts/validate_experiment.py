#!/usr/bin/env python3
"""Deterministically validate Amazing PsyCoder experiment artifacts.

This is a static preflight, not a runtime-readiness verdict. A passing result
still requires psy-exp-reviewer audit and the target-machine smoke test.
"""

from __future__ import annotations

import argparse
import ast
import csv
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    import yaml
except ImportError:  # pragma: no cover - exercised only in minimal installs
    yaml = None


PLATFORMS = {"psychopy", "jspsych", "psychtoolbox"}
INPUT_NONE = {None, "none", "null", ""}
SPECIAL_COLUMNS = {
    "subject_id", "session_id", "run_id", "timestamp", "block_name",
    "trial_number", "feedback_text", "task_name", "task_version", "date",
}
PLACEHOLDER = re.compile(r"\{([A-Za-z_][A-Za-z0-9_.-]*)\}")
MISSING = re.compile(r"\[(?:MISSING|TODO|TBD)\]", re.IGNORECASE)
CJK = re.compile(r"[\u3040-\u30ff\u3400-\u9fff\uac00-\ud7af]")
WINDOWS_ABSOLUTE = re.compile(r"^[A-Za-z]:[\\/]")
URI_SCHEME = re.compile(r"^[A-Za-z][A-Za-z0-9+.-]*://")
CORE_DATA_FIELDS = ("subject_id", "block", "trial", "condition", "timestamp")
CODE_SUFFIXES = {
    "psychopy": {".py"},
    "jspsych": {".js", ".html", ".htm"},
    "psychtoolbox": {".m"},
}
RESPONSE_EVENTS = {
    "key-down", "key-release", "click", "mouse-down", "mouse-up", "submit",
    "touch", "slider-change", "text-submit", "voice-onset", "gaze-event",
    "gaze-dwell", "hover", "drag-drop", "button",
}


@dataclass(frozen=True)
class Finding:
    level: str
    code: str
    message: str
    location: str = ""


@dataclass
class ConditionTable:
    path: Path
    headers: list[str]
    rows: list[dict[str, Any]]
    source_id: str


def add(findings: list[Finding], level: str, code: str, message: str, location: str = "") -> None:
    findings.append(Finding(level, code, message, location))


def load_config(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig")
    if yaml is not None:
        value = yaml.safe_load(text)
    else:
        try:
            value = json.loads(text)
        except json.JSONDecodeError as exc:
            raise RuntimeError("PyYAML is required for non-JSON YAML configs: pip install pyyaml") from exc
    if not isinstance(value, dict):
        raise ValueError("config root must be a mapping")
    return value


def walk(value: Any, location: str = "config") -> Iterable[tuple[str, Any]]:
    yield location, value
    if isinstance(value, dict):
        for key, child in value.items():
            yield from walk(child, f"{location}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from walk(child, f"{location}[{index}]")


def placeholders(value: Any) -> set[str]:
    found: set[str] = set()
    for _, child in walk(value):
        if isinstance(child, str):
            found.update(PLACEHOLDER.findall(child))
    return found - SPECIAL_COLUMNS


def is_response(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().lower() not in INPUT_NONE
    if value is None:
        return False
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return True


def read_condition_table(path: Path, source_id: str) -> ConditionTable:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [str(name).strip() for name in (reader.fieldnames or [])]
            rows = [{str(k).strip(): v for k, v in row.items()} for row in reader]
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise RuntimeError("openpyxl is required to validate .xlsx files: pip install openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not values:
            headers, rows = [], []
        else:
            headers = [str(value).strip() if value is not None else "" for value in values[0]]
            rows = [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
    else:
        raise ValueError(f"unsupported condition format {suffix!r}; use .csv or .xlsx")
    return ConditionTable(path, headers, rows, source_id)


def normalized(value: Any) -> str:
    return str(value if value is not None else "").strip().lower().replace("_", "-")


def truthy(value: Any) -> bool:
    return normalized(value) in {"1", "true", "yes", "y", "stop", "match", "switch", "present"}


def unsafe_project_path(value: Any) -> bool:
    raw = str(value if value is not None else "")
    path = Path(raw.replace("\\", "/"))
    return path.is_absolute() or bool(WINDOWS_ABSOLUTE.match(raw)) or bool(URI_SCHEME.match(raw)) or ".." in path.parts


def valid_duration(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return value > 0
    if isinstance(value, list) and len(value) == 2:
        low, high = value
        return (
            all(isinstance(item, (int, float)) and not isinstance(item, bool) and item > 0 for item in value)
            and low <= high
        )
    if isinstance(value, str):
        token = normalized(value)
        return bool(PLACEHOLDER.fullmatch(value.strip())) or token in {"until-key", "until-response", "self-paced"}
    return False


def check_ratio(
    findings: list[Finding],
    table: ConditionTable,
    expected: float,
    column_names: tuple[str, ...],
    predicate,
    label: str,
) -> None:
    column = next((name for name in column_names if name in table.headers), None)
    if column is None or not table.rows:
        add(findings, "error", "FID001", f"cannot verify {label}; expected one of columns {column_names}", str(table.path))
        return
    count = sum(1 for row in table.rows if predicate(row.get(column)))
    expected_count = expected * len(table.rows)
    admissible = {math.floor(expected_count), math.ceil(expected_count)}
    if count not in admissible:
        actual = count / len(table.rows)
        add(
            findings,
            "error",
            "FID002",
            f"{label} ratio is {actual:.3f}; expected {expected:.3f} "
            f"(admissible count for {len(table.rows)} rows: {sorted(admissible)})",
            str(table.path),
        )


def validate_fidelity(config: dict[str, Any], tables: list[ConditionTable], findings: list[Finding]) -> None:
    paradigm = normalized(config.get("paradigm"))
    settings = config.get("paradigm_config") if isinstance(config.get("paradigm_config"), dict) else {}
    ratios: dict[str, float] = {}
    for key in ("congruency_ratio", "go_ratio", "stop_probability", "match_ratio", "target_present_ratio", "switch_ratio"):
        if key not in settings:
            continue
        try:
            ratios[key] = float(settings[key])
        except (TypeError, ValueError):
            add(findings, "error", "FID000", f"{key} must be numeric", f"config.paradigm_config.{key}")
            continue
        if not 0 <= ratios[key] <= 1:
            add(findings, "error", "FID000", f"{key} must be in [0,1]", f"config.paradigm_config.{key}")
            ratios.pop(key)
    for table in tables:
        if "congruency_ratio" in ratios:
            check_ratio(
                findings,
                table,
                ratios["congruency_ratio"],
                ("congruency", "congruent", "condition"),
                lambda value: normalized(value) in {"congruent", "1", "true", "yes"},
                "congruency",
            )
        if "go_ratio" in ratios:
            check_ratio(
                findings,
                table,
                ratios["go_ratio"],
                ("condition", "trial_type", "is_go"),
                lambda value: normalized(value) in {"go", "1", "true", "yes"},
                "go",
            )
        if "stop_probability" in ratios:
            check_ratio(findings, table, ratios["stop_probability"], ("is_stop", "trial_type"), truthy, "stop")
        if "match_ratio" in ratios:
            check_ratio(findings, table, ratios["match_ratio"], ("is_match", "condition"), truthy, "match")
        if "target_present_ratio" in ratios:
            check_ratio(
                findings,
                table,
                ratios["target_present_ratio"],
                ("target_present", "condition"),
                truthy,
                "target-present",
            )
        if "switch_ratio" in ratios:
            check_ratio(findings, table, ratios["switch_ratio"], ("is_switch", "condition"), truthy, "switch")

        if paradigm == "stroop" and {"ink_color", "correct_response"}.issubset(table.headers):
            response_rules = config.get("response_rules") if isinstance(config.get("response_rules"), dict) else {}
            mapping = response_rules.get("mapping") if isinstance(response_rules.get("mapping"), dict) else {}
            reverse = {normalized(value): normalized(key) for key, value in mapping.items()}
            for index, row in enumerate(table.rows, 2):
                expected_key = reverse.get(normalized(row.get("ink_color")))
                actual_key = normalized(row.get("correct_response"))
                if expected_key is None:
                    add(findings, "error", "FID003", f"ink_color {row.get('ink_color')!r} is absent from response mapping", f"{table.path}:{index}")
                elif actual_key != expected_key:
                    add(
                        findings,
                        "error",
                        "FID004",
                        f"correct_response {actual_key!r} does not match ink_color; expected {expected_key!r}",
                        f"{table.path}:{index}",
                    )

        if paradigm == "go-nogo" and "max_consecutive_nogo" in settings:
            column = next((name for name in ("condition", "trial_type") if name in table.headers), None)
            if column:
                try:
                    maximum = int(settings["max_consecutive_nogo"])
                except (TypeError, ValueError):
                    add(findings, "error", "FID000", "max_consecutive_nogo must be an integer", "config.paradigm_config.max_consecutive_nogo")
                    continue
                if maximum < 0:
                    add(findings, "error", "FID000", "max_consecutive_nogo must be non-negative", "config.paradigm_config.max_consecutive_nogo")
                    continue
                run = 0
                for index, row in enumerate(table.rows, 2):
                    if normalized(row.get(column)) in {"nogo", "no-go"}:
                        run += 1
                        if run > maximum:
                            add(findings, "error", "FID005", f"more than {maximum} consecutive no-go trials", f"{table.path}:{index}")
                            break
                    else:
                        run = 0


def validate_config(config_path: Path) -> tuple[dict[str, Any], list[ConditionTable], list[Finding]]:
    findings: list[Finding] = []
    try:
        config = load_config(config_path)
    except (OSError, RuntimeError, ValueError) as exc:
        return {}, [], [Finding("error", "CFG000", str(exc), str(config_path))]

    for key in ("name", "paradigm", "platform", "runtime", "windows", "sequences", "response_rules", "randomization", "output"):
        if key not in config or config[key] in (None, "", [], {}):
            add(findings, "error", "CFG001", f"required field {key!r} is missing or empty", f"config.{key}")
    if "blocks" in config:
        add(
            findings,
            "error",
            "CFG006",
            "retired blocks configuration is not accepted; express execution with sequences and trial_sources",
            "config.blocks",
        )

    for location, value in walk(config):
        if isinstance(value, str) and MISSING.search(value):
            add(findings, "error", "CFG002", "unresolved marker found", location)

    platform = normalized(config.get("platform"))
    if platform not in PLATFORMS:
        add(findings, "error", "CFG003", f"unsupported platform {platform!r}", "config.platform")

    paradigm_file = Path(__file__).resolve().parents[1] / "psy-exp-designer" / "paradigms" / f"{normalized(config.get('paradigm'))}.md"
    if config.get("paradigm") and not paradigm_file.exists():
        add(
            findings,
            "info",
            "CFG004",
            f"no exact paradigm reference for {config.get('paradigm')!r}; validating as a custom design",
            "config.paradigm",
        )

    if not isinstance(config.get("runtime"), dict):
        add(findings, "error", "ENV000", "runtime must be a mapping", "config.runtime")
    runtime = config.get("runtime") if isinstance(config.get("runtime"), dict) else {}
    if not normalized(runtime.get("framework_version")) or normalized(runtime.get("framework_version")) in {"latest", "current"}:
        add(findings, "error", "ENV001", "runtime.framework_version must be an exact confirmed target, not latest/current", "config.runtime.framework_version")
    if normalized(runtime.get("dependency_strategy")) not in {"pinned", "lockfile"}:
        add(findings, "error", "ENV002", "runtime.dependency_strategy must be pinned or lockfile", "config.runtime.dependency_strategy")
    if not normalized(runtime.get("target_environment")):
        add(findings, "error", "ENV003", "runtime.target_environment is required", "config.runtime.target_environment")

    if not isinstance(config.get("windows"), list) or not config.get("windows"):
        add(findings, "error", "WIN000", "windows must be a non-empty list", "config.windows")
    windows = config.get("windows") if isinstance(config.get("windows"), list) else []
    window_names: list[str] = []
    response_count = 0
    for index, window in enumerate(windows):
        location = f"config.windows[{index}]"
        if not isinstance(window, dict):
            add(findings, "error", "WIN001", "window must be a mapping", location)
            continue
        for key in ("name", "content", "duration", "response"):
            if key not in window:
                add(findings, "error", "WIN002", f"required window field {key!r} is missing", f"{location}.{key}")
        if "duration" in window and not valid_duration(window.get("duration")):
            add(findings, "error", "WIN009", "duration must be positive ms, an ordered [min,max], a column placeholder, until_key/until_response, or self_paced", f"{location}.duration")
        name = str(window.get("name", "")).strip()
        if name:
            if name in window_names:
                add(findings, "error", "WIN003", f"duplicate window name {name!r}", f"{location}.name")
            window_names.append(name)
        if is_response(window.get("response")):
            response_count += 1
            event = normalized(window.get("response_event"))
            if not event:
                add(findings, "error", "WIN007", "response window requires an explicit response_event", f"{location}.response_event")
            elif event not in RESPONSE_EVENTS and not event.startswith("custom:"):
                add(findings, "error", "WIN008", f"unsupported response_event {window.get('response_event')!r}; use a documented event or custom:<name>", f"{location}.response_event")
            onset = window.get("rt_onset")
            if not onset:
                add(findings, "error", "WIN004", "response window requires rt_onset", f"{location}.rt_onset")
            elif onset != "self" and onset not in window_names and onset not in [str(w.get("name", "")) for w in windows if isinstance(w, dict)]:
                add(findings, "error", "WIN005", f"rt_onset {onset!r} does not resolve to a window", f"{location}.rt_onset")
            rationale = window.get("rt_rationale")
            if not isinstance(rationale, str) or not rationale.strip():
                add(findings, "error", "WIN010", "response window requires a non-empty rt_rationale", f"{location}.rt_rationale")
            if normalized(window.get("rt_contract_status")) != "confirmed":
                add(findings, "error", "WIN011", "response-event and RT-anchor contract must be explicitly confirmed", f"{location}.rt_contract_status")
    if windows and response_count == 0:
        add(findings, "error", "WIN006", "at least one window must accept input", "config.windows")

    if not isinstance(config.get("randomization"), dict):
        add(findings, "error", "RND000", "randomization must be a mapping", "config.randomization")
    randomization = config.get("randomization") if isinstance(config.get("randomization"), dict) else {}
    method = normalized(randomization.get("method"))
    if method not in {"random", "pseudorandom", "blocked", "counterbalanced", "fixed"}:
        add(findings, "error", "RND001", "randomization.method must be random, pseudorandom, blocked, counterbalanced, or fixed", "config.randomization.method")
    if method != "fixed" and randomization.get("seed") in (None, ""):
        add(findings, "error", "RND002", "non-fixed randomization requires a resolvable seed", "config.randomization.seed")
    if method != "fixed" and normalized(randomization.get("seed_scope")) not in {"per-session", "per-subject", "fixed"}:
        add(findings, "error", "RND003", "non-fixed randomization requires seed_scope: per_session, per_subject, or fixed", "config.randomization.seed_scope")
    if method != "fixed" and randomization.get("record_resolved_seed") is not True:
        add(findings, "error", "RND004", "record_resolved_seed must be true so realized randomization is recoverable", "config.randomization.record_resolved_seed")
    if method != "fixed" and normalized(randomization.get("seed_scope")) == "fixed" and not normalized(randomization.get("fixed_order_justification")):
        add(findings, "warning", "RND005", "a fixed seed gives every session the same pseudorandom order; document why that is intended", "config.randomization.fixed_order_justification")

    if not isinstance(config.get("output"), dict):
        add(findings, "error", "OUT000", "output must be a mapping", "config.output")
    output = config.get("output") if isinstance(config.get("output"), dict) else {}
    for key in ("directory", "filename_pattern", "incremental_save", "trial_summary"):
        if key not in output:
            add(findings, "error", "OUT001", f"output.{key} is required", f"config.output.{key}")
    if output.get("incremental_save") is not True:
        add(findings, "error", "OUT002", "output.incremental_save must be true", "config.output.incremental_save")
    if output.get("filename_pattern") and "{subject_id}" not in str(output["filename_pattern"]):
        add(findings, "error", "OUT003", "filename_pattern must contain {subject_id} to prevent overwrite", "config.output.filename_pattern")
    run_tokens = ("{session_id}", "{run_id}", "{timestamp}")
    if output.get("filename_pattern") and not any(token in str(output["filename_pattern"]) for token in run_tokens):
        add(findings, "error", "OUT005", "filename_pattern must contain {session_id}, {run_id}, or {timestamp}; a date alone is not collision-resistant", "config.output.filename_pattern")
    if output.get("trial_summary") is not True:
        add(findings, "error", "OUT004", "output.trial_summary must be true", "config.output.trial_summary")
    if output.get("directory") and unsafe_project_path(output.get("directory")):
        add(findings, "error", "OUT006", "output.directory must be a project-relative path without parent traversal", "config.output.directory")

    if not isinstance(config.get("response_rules"), dict):
        add(findings, "error", "RSP001", "response_rules must be a mapping", "config.response_rules")
    response_rules = config.get("response_rules") if isinstance(config.get("response_rules"), dict) else {}
    if config.get("paradigm_config") is not None and not isinstance(config.get("paradigm_config"), dict):
        add(findings, "error", "CFG005", "paradigm_config must be a mapping when provided", "config.paradigm_config")
    correct_refs = placeholders(response_rules.get("correct"))
    column_refs = placeholders({"windows": windows, "correct": response_rules.get("correct")})
    tables: list[ConditionTable] = []
    trial_sources = config.get("trial_sources")
    if trial_sources is None:
        trial_sources = {}
    if not isinstance(trial_sources, dict):
        add(findings, "error", "SRC001", "trial_sources must be a mapping", "config.trial_sources")
        trial_sources = {}
    for source_id, raw_path in trial_sources.items():
        location = f"config.trial_sources.{source_id}"
        if not isinstance(source_id, str) or not source_id.strip() or not isinstance(raw_path, str):
            add(findings, "error", "SRC002", "each condition-table source requires a non-empty ID and file path", location)
            continue
        condition_ref = Path(raw_path)
        condition_path = (config_path.parent / condition_ref).resolve()
        if unsafe_project_path(raw_path) or not condition_path.is_relative_to(config_path.parent.resolve()):
            add(findings, "error", "COND008", "condition file must stay inside the config project directory", location)
            continue
        if not condition_path.exists():
            add(findings, "error", "COND001", "condition file does not exist", str(condition_path))
            continue
        try:
            table = read_condition_table(condition_path, source_id)
        except (OSError, RuntimeError, ValueError) as exc:
            add(findings, "error", "COND002", str(exc), str(condition_path))
            continue
        tables.append(table)
        if not table.headers or any(not header for header in table.headers) or len(set(table.headers)) != len(table.headers):
            add(findings, "error", "COND006", "condition headers must be non-empty and unique", str(condition_path))
        if not table.rows:
            add(findings, "error", "COND007", "condition file has no data rows", str(condition_path))
        missing_columns = sorted(column_refs - set(table.headers))
        if missing_columns:
            add(findings, "error", "COND003", f"missing referenced columns: {', '.join(missing_columns)}", str(condition_path))
        for column in correct_refs & set(table.headers):
            for row_number, row in enumerate(table.rows, 2):
                if row.get(column) in (None, ""):
                    add(findings, "error", "COND009", f"correct-response column {column!r} contains an empty value; encode an explicit no-response rule when intended", f"{table.path}:{row_number}")

    sequences = config.get("sequences") if isinstance(config.get("sequences"), list) else []
    for index, seq in enumerate(sequences):
        loc = f"config.sequences[{index}]"
        if not isinstance(seq, dict):
            add(findings, "error", "SEQ001", "sequence must be a mapping", loc)
            continue
        if not str(seq.get("name", "")).strip():
            add(findings, "error", "SEQ002", "sequence name is required", f"{loc}.name")
        if not isinstance(seq.get("window_ids"), list) or not seq.get("window_ids"):
            add(findings, "error", "SEQ003", "window_ids must be a non-empty list", f"{loc}.window_ids")
        else:
            for wid in seq["window_ids"]:
                if wid not in window_names:
                    add(findings, "error", "SEQ004", f"window_ids references unknown window {wid!r}", f"{loc}.window_ids")
        source_id = seq.get("trial_source_id")
        if source_id is not None and source_id not in trial_sources:
            add(findings, "error", "SEQ014", f"trial_source_id references unknown source {source_id!r}", f"{loc}.trial_source_id")
        execution = seq.get("execution")
        if not isinstance(execution, dict):
            add(findings, "error", "SEQ005", "execution must be a mapping", f"{loc}.execution")
        else:
            if "mode" in execution:
                add(findings, "error", "SEQ006", "retired execution.mode is not accepted", f"{loc}.execution.mode")
            reps = execution.get("repetitions")
            if not isinstance(reps, int) or isinstance(reps, bool) or not 1 <= reps <= 10000:
                add(findings, "error", "SEQ007", "execution.repetitions must be an integer from 1 to 10000", f"{loc}.execution.repetitions")
            order_mode = execution.get("order_mode")
            if order_mode not in {"table_order", "fixed_random", "fully_random"}:
                add(findings, "error", "SEQ008", "execution.order_mode must be table_order, fixed_random, or fully_random", f"{loc}.execution.order_mode")
            if order_mode == "fixed_random" and not str(execution.get("seed", "")).strip():
                add(findings, "error", "SEQ015", "fixed_random requires a non-empty seed", f"{loc}.execution.seed")
            reshuffle = execution.get("reshuffle_each_cycle")
            if reshuffle is not None and not isinstance(reshuffle, bool):
                add(findings, "error", "SEQ016", "reshuffle_each_cycle must be boolean", f"{loc}.execution.reshuffle_each_cycle")
        show_in = seq.get("show_in")
        if show_in is not None:
            if not isinstance(show_in, list):
                add(findings, "error", "SEQ009", "show_in must be a list", f"{loc}.show_in")
            else:
                valid_contexts = {"practice", "formal", "rest", "debrief"}
                for context in show_in:
                    if context not in valid_contexts:
                        add(findings, "error", "SEQ010", f"invalid show_in value {context!r}; use practice/formal/rest/debrief", f"{loc}.show_in")

    timing_overrides = config.get("timing") if isinstance(config.get("timing"), dict) else {}
    if timing_overrides.get("iti") and any(w.get("name", "").lower() == "iti" for w in windows):
        add(findings, "warning", "SEQ012", "both timing.iti override and an ITI window exist — ITI window takes precedence", "config.timing.iti")
    if tables:
        all_rows = [tuple(row.values()) for table in tables for row in table.rows]
        if len(set(all_rows)) <= 1:
            add(findings, "warning", "SEQ013", "all condition rows are identical — experiment has no variation", "config.trial_sources")

    stimulus_folder = config.get("stimulus_folder")
    if stimulus_folder:
        folder_ref = Path(str(stimulus_folder))
        folder = (config_path.parent / folder_ref).resolve()
        if unsafe_project_path(stimulus_folder) or not folder.is_relative_to(config_path.parent.resolve()):
            add(findings, "error", "STIM003", "stimulus_folder must stay inside the config project directory", "config.stimulus_folder")
        elif not folder.is_dir():
            add(findings, "error", "STIM001", "stimulus_folder does not exist", str(folder))
        else:
            content_refs = placeholders([window.get("content") for window in windows if isinstance(window, dict)])
            for table in tables:
                for column in content_refs & set(table.headers):
                    for index, row in enumerate(table.rows, 2):
                        value = row.get(column)
                        if value not in (None, "") and not (folder / str(value)).exists():
                            add(findings, "error", "STIM002", f"stimulus file {value!r} does not exist", f"{table.path}:{index}")

    validate_fidelity(config, tables, findings)
    return config, tables, findings


CODE_RULES: dict[str, dict[str, list[tuple[str, str]]]] = {
    "psychopy": {
        "required": [
            (r"from\s+psychopy", "PsychoPy imports"),
            (r"Keyboard\s*\([^)]*backend\s*=", "explicit keyboard backend"),
            (r"callOnFlip\s*\(", "flip-synchronized callback"),
            (r"clearEvents", "pre-stimulus keyboard clear"),
            (r"waitRelease\s*=\s*False", "press-time response collection"),
            (r"\.rt\b", "hardware-timestamped RT"),
            (r"\brandom\.(?:seed|Random)\s*\(|\b(?:np|numpy)\.random\.default_rng\s*\(|\bTrialHandler\s*\([^)]*\bseed\s*=", "reproducible randomization seed"),
            (r"\btry\s*:", "protected experiment loop"),
            (r"\bfinally\s*:", "cleanup block"),
            (r"\.flush\s*\(|\bos\.fsync\s*\(|\.commit\s*\(", "per-trial durable persistence"),
            (r"escape", "emergency quit"),
        ],
        "forbidden": [
            (r"\btime\.sleep\s*\(", "time.sleep blocks the event loop"),
            (r"event\.(?:getKeys|waitKeys)\s*\([^)]*maxWait", "blocking legacy keyboard API"),
            (r"\brt\s*=\s*(?:\([^\n]*\)\s*)?time\.time\s*\(", "manual RT clock"),
            (r"waitRelease\s*=\s*True", "release-time response collection"),
            (r"\bexec\s*\(", "dynamic condition execution"),
            (r"\bglobals\s*\(\s*\)\s*\[", "condition injection into global namespace"),
        ],
        "warnings": [
            (r"\btime\.time\s*\(", "time.time() requires semantic review: allow wall-clock metadata only, never keyboard RT"),
        ],
    },
    "jspsych": {
        "required": [
            (r"\binitJsPsych\s*\(", "jsPsych initialization"),
            (r"\bjsPsych\.run\s*\(", "timeline launch"),
            (r"\bjsPsychPreload\b", "preload plugin"),
            (r"randomization\.setSeed\s*\(", "reproducible randomization seed"),
            (r"pluginAPI\.compareKeys\s*\(", "normalized key comparison"),
            (r"\.localSave\s*\(|XMLHttpRequest|fetch\s*\(", "final durable export"),
            (r"\b(?:on_data_update|on_trial_finish)\s*(?::|\()", "per-trial persistence callback"),
            (r"localStorage|indexedDB|XMLHttpRequest|fetch\s*\(", "durable per-trial checkpoint"),
            (r"addEventListener\s*\(\s*['\"]keydown['\"]", "centralized keyboard abort listener"),
            (r"removeEventListener\s*\(\s*['\"]keydown['\"]", "abort listener cleanup"),
            (r"\babortExperiment\s*\(", "jsPsych emergency abort"),
            (r"escape", "emergency quit key"),
        ],
        "forbidden": [
            (r"\bjsPsych\.init\s*\(", "removed jsPsych 6 initialization"),
            (r"\btype\s*:\s*['\"](?:html|image|audio|video|categorize|survey|instructions|fullscreen|preload|call-function)[A-Za-z0-9_-]*['\"]", "string plugin type"),
            (r"timelineVariable\s*\([^)]*,\s*true\s*\)", "legacy jsPsych 7 immediate timeline-variable evaluation; use evaluateTimelineVariable()"),
            (r"\b(?:rt|response_time)\s*[:=][^;\n]*\bDate\.now\s*\(", "manual RT clock"),
            (r"\bchoices\s*:\s*\[[^\]]*['\"]escape['\"]", "Escape must not be mixed into scored choices; use the centralized abort listener"),
        ],
        "warnings": [
            (r"\bDate\.now\s*\(", "Date.now() requires semantic review: allow wall-clock metadata/run IDs only, never scored RT"),
            (r"\bMath\.random\s*\(", "Math.random() is seeded only after jsPsych.randomization.setSeed() executes; prefer an explicit jsPsych.randomization API and record the resolved seed"),
            (r"\bset(?:Timeout|Interval)\s*\(", "manual timer requires semantic review: do not replace jsPsych trial timing APIs"),
        ],
    },
    "psychtoolbox": {
        "required": [
            (r"PsychImaging", "Psychtoolbox display setup"),
            (r"SkipSyncTests['\"]?\s*,\s*0", "production sync tests"),
            (r"KbQueueCreate", "keyboard queue creation"),
            (r"KbQueueStart", "keyboard queue start"),
            (r"KbQueueFlush", "per-trial keyboard flush"),
            (r"KbQueueCheck", "timestamped response collection"),
            (r"firstPress", "hardware key timestamp"),
            (r"Screen\s*\(\s*['\"]Flip['\"]", "VBL-synchronized flip"),
            (r"\brng\s*\([^)]*['\"]twister['\"]", "reproducible randomization seed"),
            (r"\bfopen\s*\([^\n;]*['\"]a['\"]|\bwritetable\s*\(|\bsave\s*\([^\n;]*-append|\bmatfile\s*\(", "per-trial durable persistence"),
            (r"\btry\b", "protected experiment loop"),
            (r"\bcatch\b", "cleanup on error"),
            (r"Priority\s*\(\s*0\s*\)", "priority restoration"),
            (r"ShowCursor", "cursor restoration"),
        ],
        "forbidden": [
            (r"SkipSyncTests['\"]?\s*,\s*1", "disabled sync tests"),
            (r"\brt\s*=\s*GetSecs", "manual RT clock"),
            (r"\brng\s*\(\s*['\"]shuffle['\"]", "non-reproducible random seed"),
        ],
        "warnings": [
            (r"\bKbWait\b", "KbWait is blocking; verify it appears only on a justified static, non-critical screen"),
            (r"\bWaitSecs\s*\(", "WaitSecs is blocking; verify it is not used where flips, triggers, input, or abort handling must continue"),
        ],
    },
}


def javascript_source(path: Path, text: str) -> str:
    if path.suffix.lower() not in {".html", ".htm"}:
        return text
    return "\n".join(
        match.group(1)
        for match in re.finditer(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", text, flags=re.IGNORECASE | re.DOTALL)
    )


def python_string_payload(text: str) -> str:
    """Return Python string literals while excluding comments and identifiers."""
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return ""
    docstrings: set[int] = set()
    for owner in ast.walk(tree):
        if isinstance(owner, (ast.Module, ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)) and owner.body:
            first = owner.body[0]
            if (
                isinstance(first, ast.Expr)
                and isinstance(first.value, ast.Constant)
                and isinstance(first.value.value, str)
            ):
                docstrings.add(id(first.value))
    return "\n".join(
        node.value
        for node in ast.walk(tree)
        if isinstance(node, ast.Constant) and isinstance(node.value, str) and id(node) not in docstrings
    )


def code_traits(config: dict[str, Any] | None) -> dict[str, bool]:
    if config is None:
        return {"keyboard": True, "key_release": False, "rt": True, "scoring": True, "randomized": True, "media": True, "response": True}
    windows = config.get("windows") if isinstance(config.get("windows"), list) else []
    responses = [window.get("response") for window in windows if isinstance(window, dict) and is_response(window.get("response"))]
    response_text = " ".join(str(value).lower() for value in responses)
    non_keyboard_tokens = ("mouse", "click", "slider", "rating", "touch", "voice", "text", "gaze")
    keyboard = bool(responses) and (
        any(isinstance(value, (list, tuple, set)) for value in responses)
        or not any(token in response_text for token in non_keyboard_tokens)
    )
    key_release = any(
        normalized(window.get("response_event")) == "key-release"
        for window in windows
        if isinstance(window, dict)
    )
    rt = any(
        isinstance(window, dict)
        and is_response(window.get("response"))
        and (window.get("rt_onset") or "rt" in (window.get("data") or []))
        for window in windows
    )
    response_rules = config.get("response_rules") if isinstance(config.get("response_rules"), dict) else {}
    randomization = config.get("randomization") if isinstance(config.get("randomization"), dict) else {}
    scoring = bool(response_rules.get("correct"))
    randomized = normalized(randomization.get("method")) != "fixed"
    media_pattern = re.compile(r"\.(?:png|jpe?g|gif|webp|wav|mp3|ogg|mp4|webm)\b", re.IGNORECASE)
    media = bool(config.get("stimulus_folder")) or any(
        isinstance(window, dict) and media_pattern.search(str(window.get("content", "")))
        for window in windows
    )
    return {"keyboard": keyboard, "key_release": key_release, "rt": rt, "scoring": scoring, "randomized": randomized, "media": media, "response": bool(responses)}


def skipped_requirements(platform: str, traits: dict[str, bool]) -> set[str]:
    skipped: set[str] = set()
    if not traits["randomized"]:
        skipped.add("reproducible randomization seed")
    if platform == "psychopy":
        if not traits["keyboard"]:
            skipped.update({"explicit keyboard backend", "pre-stimulus keyboard clear", "press-time response collection", "hardware-timestamped RT"})
        if traits["key_release"]:
            skipped.add("press-time response collection")
        if not traits["rt"]:
            skipped.add("flip-synchronized callback")
    elif platform == "jspsych":
        if not (traits["keyboard"] and traits["scoring"]):
            skipped.add("normalized key comparison")
        if not traits["media"]:
            skipped.add("preload plugin")
    elif platform == "psychtoolbox" and not traits["keyboard"]:
        skipped.update({"keyboard queue creation", "keyboard queue start", "per-trial keyboard flush", "timestamped response collection", "hardware key timestamp"})
    return skipped


def validate_code(path: Path, platform: str, config: dict[str, Any] | None = None) -> list[Finding]:
    findings: list[Finding] = []
    try:
        text = path.read_text(encoding="utf-8-sig")
    except OSError as exc:
        return [Finding("error", "CODE000", str(exc), str(path))]
    if MISSING.search(text):
        add(findings, "error", "CODE001", "generated code contains unresolved marker", str(path))
    if path.suffix.lower() not in CODE_SUFFIXES.get(platform, set()):
        expected = ", ".join(sorted(CODE_SUFFIXES.get(platform, set())))
        add(findings, "error", "CODE008", f"code suffix does not match {platform}; expected one of: {expected}", str(path))

    if platform == "psychopy":
        try:
            ast.parse(text, filename=str(path))
        except SyntaxError as exc:
            add(findings, "error", "CODE002", f"Python syntax error: {exc.msg}", f"{path}:{exc.lineno}")
    elif platform == "jspsych":
        source = javascript_source(path, text)
        if not source.strip():
            add(findings, "error", "CODE003", "no inline JavaScript found", str(path))
        elif shutil.which("node"):
            with tempfile.NamedTemporaryFile("w", suffix=".js", encoding="utf-8", delete=False) as handle:
                handle.write(source)
                temp_path = Path(handle.name)
            try:
                result = subprocess.run(["node", "--check", str(temp_path)], capture_output=True, text=True, check=False)
            finally:
                temp_path.unlink(missing_ok=True)
            if result.returncode:
                detail = (result.stderr or result.stdout).strip().splitlines()[-1]
                add(findings, "error", "CODE004", f"JavaScript syntax check failed: {detail}", str(path))
        else:
            add(findings, "warning", "CODE005", "Node.js unavailable; JavaScript syntax was not checked", str(path))
    elif platform == "psychtoolbox":
        add(findings, "warning", "CODE006", "MATLAB syntax/runtime requires target-machine checkcode and smoke test", str(path))

    if platform == "psychopy" and CJK.search(python_string_payload(text)) and not re.search(r"FONT_AUTO_DETECT|MANUAL_FONT_PATH|FONT_CONFIG", text):
        add(findings, "error", "CODE007", "CJK text requires an explicit font configuration block", str(path))

    traits = code_traits(config)
    skip = skipped_requirements(platform, traits)
    rules = CODE_RULES.get(platform, {})
    for pattern, description in rules.get("required", []):
        if description in skip:
            continue
        if not re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL):
            add(findings, "error", "CODE100", f"missing required pattern: {description}", str(path))
    for pattern, description in rules.get("forbidden", []):
        if platform == "psychopy" and traits["key_release"] and description == "release-time response collection":
            continue
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            line = text.count("\n", 0, match.start()) + 1
            add(findings, "error", "CODE200", f"forbidden pattern: {description}", f"{path}:{line}")
    if platform == "psychopy" and traits["keyboard"] and traits["key_release"] and not re.search(r"waitRelease\s*=\s*True", text):
        add(findings, "error", "CODE102", "key_release response_event requires waitRelease=True", str(path))
    if platform == "psychopy" and traits["keyboard"] and traits["key_release"] and not re.search(r"\.duration\b|\btUp\b", text):
        add(findings, "error", "CODE103", "key_release scoring/timing must use the release duration/timestamp, not key-down .rt alone", str(path))
    for pattern, description in rules.get("warnings", []):
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            line = text.count("\n", 0, match.start()) + 1
            add(findings, "warning", "CODE300", description, f"{path}:{line}")
    required_fields = list(CORE_DATA_FIELDS)
    if traits["response"]:
        required_fields.append("response")
    if traits["rt"]:
        required_fields.append("rt")
    if traits["scoring"]:
        required_fields.extend(("correct_response", "accuracy"))
    missing_fields = [field for field in required_fields if not re.search(rf"\b{re.escape(field)}\b", text)]
    if missing_fields:
        add(
            findings,
            "error",
            "CODE101",
            f"missing output fields required by this config: {', '.join(missing_fields)}",
            str(path),
        )
    return findings


def serialize(config_path: Path, code_path: Path | None, config: dict[str, Any], tables: list[ConditionTable], findings: list[Finding]) -> dict[str, Any]:
    errors = sum(finding.level == "error" for finding in findings)
    warnings = sum(finding.level == "warning" for finding in findings)
    return {
        "config": str(config_path),
        "code": str(code_path) if code_path else None,
        "platform": normalized(config.get("platform")),
        "condition_files": [str(table.path) for table in tables],
        "static_gate_passed": errors == 0,
        "pre_code_ready": errors == 0 and code_path is None,
        "code_static_gate_passed": errors == 0 if code_path is not None else None,
        "ready_for_collection": False,
        "runtime_smoke_test_required": True,
        "errors": errors,
        "warnings": warnings,
        "findings": [asdict(finding) for finding in findings],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("config", type=Path, help="experiment config YAML/JSON")
    parser.add_argument("--code", type=Path, help="generated .py/.js/.html/.m artifact")
    parser.add_argument("--json", action="store_true", help="emit machine-readable JSON")
    args = parser.parse_args()

    config_path = args.config.resolve()
    config, tables, findings = validate_config(config_path)
    platform = normalized(config.get("platform"))
    if args.code and platform in PLATFORMS:
        findings.extend(validate_code(args.code.resolve(), platform, config))
    report = serialize(config_path, args.code.resolve() if args.code else None, config, tables, findings)

    if args.json:
        print(json.dumps(report, indent=2, ensure_ascii=False))
    else:
        for finding in findings:
            location = f" ({finding.location})" if finding.location else ""
            print(f"{finding.level.upper()} {finding.code}: {finding.message}{location}")
        status = "PASS" if report["static_gate_passed"] else "FAIL"
        print(f"Static experiment validation: {status} ({report['errors']} errors, {report['warnings']} warnings)")
        if report["pre_code_ready"]:
            print("The validated config is pre-code-ready; no implementation was reviewed.")
        elif report["code_static_gate_passed"]:
            print("The provided config and code passed deterministic static checks.")
        print("Runtime readiness is not inferred; psy-exp-reviewer audit and target-machine smoke tests remain mandatory.")
    return 0 if report["static_gate_passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
